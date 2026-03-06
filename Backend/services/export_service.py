import io
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill
from lxml import etree

logger = logging.getLogger(__name__)

# ─── Field helpers (handle both snake_case and camelCase) ─────────────────────

def _epic_title(e: dict) -> str:
    return e.get("epic_name") or e.get("title") or ""

def _feature_title(f: dict) -> str:
    return f.get("feature_name") or f.get("title") or ""

def _uc_title(u: dict) -> str:
    return u.get("use_case_title") or u.get("title") or ""

def _tc_title(t: dict) -> str:
    return t.get("test_case_title") or t.get("title") or ""

def _tc_id(t: dict) -> str:
    return t.get("test_case_id") or t.get("id") or ""

def _steps(t: dict) -> list:
    s = t.get("test_steps") or t.get("steps") or []
    return s if isinstance(s, list) else [s]

def _expected(t: dict) -> str:
    return t.get("expected_result") or t.get("expectedResult") or ""

def _compliance(t: dict) -> list:
    c = t.get("compliance_mapping") or t.get("complianceMapping") or []
    return c if isinstance(c, list) else [c]

def _use_cases(f: dict) -> list:
    return f.get("use_cases") or f.get("useCases") or []

def _test_cases(u: dict) -> list:
    return u.get("test_cases") or u.get("testCases") or []

def _jira_key(item: dict) -> str:
    return item.get("jira_issue_key") or item.get("jiraIssueKey") or ""


def export_to_excel(artifacts: dict) -> bytes:
    """Export project artifacts to Excel workbook bytes."""
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")

    # ─── Summary sheet (unchanged) ──────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Project", artifacts.get("projectName", "")])
    ws_summary.append(["Jira Key", artifacts.get("jiraProjectKey", "")])
    ws_summary.append([])
    ws_summary.append(["Metric", "Count"])
    for cell in ws_summary[4]:
        cell.font = header_font
        cell.fill = header_fill
    ws_summary.append(["Epics",      artifacts.get("totalEpics", 0)])
    ws_summary.append(["Features",   artifacts.get("totalFeatures", 0)])
    ws_summary.append(["Use Cases",  artifacts.get("totalUseCases", 0)])
    ws_summary.append(["Test Cases", artifacts.get("totalTestCases", 0)])

    # ─── Test Cases sheet ───────────────────────────────────────────────────────
    ws_tc = wb.create_sheet("Test Cases")
    headers = [
        "Epic ID", "Epic Title", "Epic Description", "Epic Jira Issue Key",
        "Feature ID", "Feature Title", "Feature Description", "Feature Jira Issue Key",
        "Use Case ID", "Use Case Title", "Use Case Description", "Use Case Jira Issue Key",
        "Acceptance Criteria", "Test Scenarios Outline",
        "Test Case ID", "Test Case Title", "Preconditions", "Test Type", "Test Steps",
        "Expected Result", "Compliance Mapping", "Test Case Jira Issue Key",
        "Priority", "Comments", "Model Explanation", "Review Status", "Jira Project Key",
    ]
    ws_tc.append(headers)
    for cell in ws_tc[1]:
        cell.font = header_font
        cell.fill = header_fill

    project_jira_key = artifacts.get("jiraProjectKey") or artifacts.get("jira_project_key") or ""

    for epic in artifacts.get("epics", []):
        epic_id    = epic.get("epic_id", "")
        epic_title = _epic_title(epic)
        epic_desc  = epic.get("description", "")
        epic_jira  = _jira_key(epic)
        epic_proj_key = project_jira_key

        for feature in epic.get("features", []):
            feat_id    = feature.get("feature_id", "")
            feat_title = _feature_title(feature)
            feat_desc  = feature.get("description", "")
            feat_jira  = _jira_key(feature)

            for uc in _use_cases(feature):
                uc_id    = uc.get("use_case_id", "")
                uc_title = _uc_title(uc)
                uc_desc  = uc.get("description", "")
                uc_jira  = _jira_key(uc)

                ac_raw = uc.get("acceptance_criteria") or []
                ac = " | ".join(ac_raw) if isinstance(ac_raw, list) else str(ac_raw or "")

                ts_raw = uc.get("test_scenarios_outline") or []
                ts = " | ".join(ts_raw) if isinstance(ts_raw, list) else str(ts_raw or "")

                test_cases = _test_cases(uc)
                if test_cases:
                    for tc in test_cases:
                        preconds_raw = tc.get("preconditions") or tc.get("pre_conditions") or []
                        preconds = " | ".join(preconds_raw) if isinstance(preconds_raw, list) else str(preconds_raw or "")

                        steps_raw = _steps(tc)
                        steps_text = " | ".join(steps_raw) if steps_raw else ""

                        comp_raw = _compliance(tc)
                        compliance_text = " | ".join(comp_raw) if comp_raw else ""

                        ws_tc.append([
                            epic_id, epic_title, epic_desc, epic_jira,
                            feat_id, feat_title, feat_desc, feat_jira,
                            uc_id, uc_title, uc_desc, uc_jira, ac, ts,
                            _tc_id(tc),
                            _tc_title(tc),
                            preconds,
                            tc.get("test_type") or tc.get("testType") or "Functional",
                            steps_text,
                            _expected(tc),
                            compliance_text,
                            _jira_key(tc),
                            tc.get("priority", ""),
                            tc.get("comments", ""),
                            tc.get("model_explanation") or tc.get("modelExplanation") or "",
                            tc.get("review_status") or tc.get("reviewStatus") or "",
                            epic_proj_key,
                        ])
                else:
                    # Use case with no test cases — output a row with empty TC fields
                    ws_tc.append([
                        epic_id, epic_title, epic_desc, epic_jira,
                        feat_id, feat_title, feat_desc, feat_jira,
                        uc_id, uc_title, uc_desc, uc_jira, ac, ts,
                        "", "", "", "", "", "", "", "", "", "", "", "", epic_proj_key,
                    ])

    for ws in [ws_summary, ws_tc]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_to_xml(artifacts: dict) -> bytes:
    """Export project artifacts to XML bytes (mirrors the frontend exportArtifactsToXml)."""
    jira_key = artifacts.get("jiraProjectKey") or artifacts.get("jira_project_key") or ""

    root = etree.Element("project")
    _t(root, "project_id",      artifacts.get("projectId", ""))
    _t(root, "name",            artifacts.get("projectName", ""))
    _t(root, "jira_project_key", jira_key)

    stats = etree.SubElement(root, "statistics")
    _t(stats, "total_epics",      str(artifacts.get("totalEpics", 0)))
    _t(stats, "total_features",   str(artifacts.get("totalFeatures", 0)))
    _t(stats, "total_use_cases",  str(artifacts.get("totalUseCases", 0)))
    _t(stats, "total_test_cases", str(artifacts.get("totalTestCases", 0)))

    epics_el = etree.SubElement(root, "epics")
    for epic in artifacts.get("epics", []):
        e_el = etree.SubElement(epics_el, "epic")
        _t(e_el, "epic_id",         epic.get("epic_id", ""))
        _t(e_el, "title",           _epic_title(epic))
        _t(e_el, "description",     epic.get("description", ""))
        _t(e_el, "priority",        epic.get("priority", "Medium"))
        _t(e_el, "jira_issue_key",  _jira_key(epic))
        _t(e_el, "jira_project_key", jira_key)

        feats_el = etree.SubElement(e_el, "features")
        for feat in epic.get("features", []):
            f_el = etree.SubElement(feats_el, "feature")
            _t(f_el, "feature_id",      feat.get("feature_id", ""))
            _t(f_el, "title",           _feature_title(feat))
            _t(f_el, "description",     feat.get("description", ""))
            _t(f_el, "priority",        feat.get("priority") or epic.get("priority", "Medium"))
            _t(f_el, "jira_issue_key",  _jira_key(feat))
            _t(f_el, "jira_project_key", jira_key)

            ucs_el = etree.SubElement(f_el, "use_cases")
            for uc in _use_cases(feat):
                uc_el = etree.SubElement(ucs_el, "use_case")
                _t(uc_el, "use_case_id",    uc.get("use_case_id", ""))
                _t(uc_el, "title",          _uc_title(uc))
                _t(uc_el, "description",    uc.get("description", ""))
                _t(uc_el, "priority",       uc.get("priority") or feat.get("priority") or epic.get("priority", "Medium"))
                _t(uc_el, "jira_issue_key", _jira_key(uc))
                _t(uc_el, "jira_project_key", jira_key)
                _t(uc_el, "review_status",  uc.get("review_status") or uc.get("reviewStatus") or "")
                if uc.get("model_explanation") or uc.get("modelExplanation"):
                    _t(uc_el, "model_explanation", uc.get("model_explanation") or uc.get("modelExplanation", ""))

                tcs_el = etree.SubElement(uc_el, "test_cases")
                for tc in _test_cases(uc):
                    tc_el = etree.SubElement(tcs_el, "test_case")
                    _t(tc_el, "test_case_id",  _tc_id(tc))
                    _t(tc_el, "title",         _tc_title(tc))
                    _t(tc_el, "priority",      tc.get("priority", "Medium"))
                    _t(tc_el, "test_type",     tc.get("test_type") or tc.get("testType") or "Functional")
                    _t(tc_el, "review_status", tc.get("review_status") or tc.get("reviewStatus") or "Pending")
                    _t(tc_el, "jira_issue_key", _jira_key(tc))
                    _t(tc_el, "jira_project_key", jira_key)

                    preconds = tc.get("preconditions") or tc.get("pre_conditions")
                    if preconds:
                        pc_el = etree.SubElement(tc_el, "preconditions")
                        items = preconds if isinstance(preconds, list) else [preconds]
                        for p in items:
                            _t(pc_el, "precondition", p)

                    steps = _steps(tc)
                    if steps:
                        steps_el = etree.SubElement(tc_el, "test_steps")
                        for i, s in enumerate(steps, 1):
                            s_el = etree.SubElement(steps_el, "step")
                            s_el.set("step_number", str(i))
                            s_el.text = s

                    if _expected(tc):
                        _t(tc_el, "expected_result", _expected(tc))
                    if tc.get("model_explanation") or tc.get("modelExplanation"):
                        _t(tc_el, "model_explanation", tc.get("model_explanation") or tc.get("modelExplanation", ""))
                    if tc.get("comments"):
                        _t(tc_el, "comments", tc["comments"])

                    compliance = _compliance(tc)
                    if compliance:
                        cm_el = etree.SubElement(tc_el, "compliance_mapping")
                        for c in compliance:
                            _t(cm_el, "compliance", c)

    return etree.tostring(root, pretty_print=True, xml_declaration=True, encoding="UTF-8")


def _t(parent: etree._Element, tag: str, text: str) -> etree._Element:
    """Create a sub-element with text content."""
    el = etree.SubElement(parent, tag)
    el.text = str(text) if text is not None else ""
    return el

